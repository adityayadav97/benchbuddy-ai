"""Knowledge-base loader.

Reads the PMO FAQ workbooks shipped with the kata into a unified, de-duplicated
list of FAQ rows. We merge the 10-row baseline KB with the 50-row extended KB
so that the agent always has the richer signal available.
"""

from __future__ import annotations

import os
from dataclasses import dataclass
from typing import List

import openpyxl


@dataclass(frozen=True)
class FAQRow:
    category: str
    question: str
    answer: str

    @property
    def text(self) -> str:
        return f"{self.category}. {self.question} {self.answer}"


def _read_workbook(path: str) -> List[FAQRow]:
    if not os.path.exists(path):
        return []
    wb = openpyxl.load_workbook(path, data_only=True)
    rows: List[FAQRow] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        header = None
        for raw in ws.iter_rows(values_only=True):
            if not any(raw):
                continue
            if header is None:
                header = [str(c).strip().lower() if c else "" for c in raw]
                continue
            record = {header[i]: raw[i] for i in range(min(len(header), len(raw)))}
            cat = (record.get("category") or "General").strip()
            q = (record.get("question") or "").strip()
            a = (record.get("approved answer") or record.get("answer") or "").strip()
            if q and a:
                rows.append(FAQRow(cat, q, a))
    return rows


def load_knowledge_base(data_dir: str) -> List[FAQRow]:
    """Load and merge all KB files under *data_dir*. De-duplicates on (category, question)."""

    candidates = [
        "PMO_FAQ_Knowledge_Base_50_FAQs.xlsx",
        "PMO_FAQ_Knowledge_Base.xlsx",
    ]
    seen = set()
    merged: List[FAQRow] = []
    for name in candidates:
        for row in _read_workbook(os.path.join(data_dir, name)):
            key = (row.category.lower(), row.question.lower())
            if key in seen:
                continue
            seen.add(key)
            merged.append(row)
    return merged
